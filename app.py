import openpyxl as xl
from openpyxl.chart import BarChart, Reference

wb = xl.load_workbook("transactions.xlsx")
sheet = wb["Sheet1"]

for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    if isinstance(cell.value, (int, float)):
        corrected_price = cell.value * 0.9
        sheet.cell(row, 4).value = corrected_price

values = Reference(
    sheet,
    min_row=2,
    max_row=sheet.max_row,
    min_col=4,
    max_col=4
)
chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart, "F2")

wb.save("transactions2.xlsx")

# # weight converter
# weight = int(input("weight: "))
# unit = input("(L)bs or (K)g: ")

# if unit.lower() == "l":
#     converted = round(weight * 0.45)
#     print(f"you are {converted} kilos")
# else:
#     converted = round(weight / 0.45)
#     print(f"you are {converted} pounds")

# # guessing game with max of 4 trials
# num_of_trials = 0

# while num_of_trials < 4:
#     guess = int(input("guess: "))
#     if guess == 9:
#         print("you won!")
#         break
#     num_of_trials += 1
# else:
#     print("you lost!")
